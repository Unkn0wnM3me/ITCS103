import openpyxl as op
import datetime

current_year = datetime.datetime.now().year

wbk = op.Workbook()
sheet = wbk.active

sheet.append(["ID", "First Name", "Last Name", "Birth Year", "Age"])

print("=========================================")
print("        Favorite People Recorder")
print("=========================================\n")

for i in range(1, 4):
    print(f"Enter details for Person {i}:")
    first_name = input("First Name: ")
    last_name = input("Last Name: ")
    birth_year = int(input("Birth Year: "))
    
    age = current_year - birth_year
    
    person_id = i
    
    sheet.append([person_id, first_name, last_name, birth_year, age])
    print()  # Just for spacing

filename = "favorite_people.xlsx"
wbk.save(filename)
print("-----------------------------------------")
print(f"Success: Data saved to {filename}!")
print("-----------------------------------------\n")

print("Saved Records in Excel File:")
print("-----------------------------------------")

read_wbk = op.load_workbook(filename)
read_sheet = read_wbk.active

for rows in read_sheet.iter_rows(values_only=True):
    print(rows)