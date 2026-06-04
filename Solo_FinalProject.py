import tkinter as tk
from tkinter import ttk, messagebox
import openpyxl as op
import os

EXCEL_FILE = "Solo_Database.xlsx"

if not os.path.exists(EXCEL_FILE):
    workbook = op.Workbook()
    sheet = workbook.active
    sheet.title = "Customers"
    
    sheet["A1"] = "Customer ID"
    sheet["B1"] = "Last Name"
    sheet["C1"] = "First Name"
    sheet["D1"] = "Middle Name"
    sheet["E1"] = "Birth Year"
    sheet["F1"] = "Age"
    
    sheet.append(["1", "Solo", "Ken Gimelson", "Bernal", 2007, 19])
    workbook.save(EXCEL_FILE)


def validate_input():
    first = fname_entry.get().strip()
    last = lname_entry.get().strip()
    by = birth_entry.get().strip()
    
    if not first or not last or not by:
        messagebox.showerror("Error", "Forst name, Last name, and Birth year are required.")
        return False
    if not by.isdigit():
        messagebox.showerror("Error", "Birth year mustbe a number.")
        return False
    return True

def display_excel():
    workbook = op.load_workbook(EXCEL_FILE)
    sheet = workbook.active
    
    for row in tree.get_children():
        tree.delete(row)
    
    for row in sheet.iter_rows(min_row=2, values_only=True):
        tree.insert("", tk.END, values=row)

def append_excel():
    if validate_input():
        first = fname_entry.get().strip()
        middle = mname_entry.get().strip()
        last = lname_entry.get().strip()
        by = int(birth_entry.get().strip())
        age = 2026 - by  
        
        workbook = op.load_workbook(EXCEL_FILE)
        sheet = workbook.active
        
        
        new_id = sheet.max_row
        
        sheet.append([new_id, last, first, middle, by, age])
        workbook.save(EXCEL_FILE)
        
        messagebox.showinfo("Successs!", "Record added successfully!!")
        display_excel()
        clear_entries()


def select_record(event):
    selected = tree.focus()
    values = tree.item(selected, "values")
    if values:
        
        lname_entry.delete(0, tk.END)
        fname_entry.delete(0, tk.END)
        mname_entry.delete(0, tk.END)
        birth_entry.delete(0, tk.END)
        
        
        lname_entry.insert(0, values[1])
        fname_entry.insert(0, values[2])
        mname_entry.insert(0, values[3])
        birth_entry.insert(0, values[4])


def update_data():
    selected = tree.focus()
    if not selected:
        messagebox.showerror("Error", "Select a record first!!")
        return
        
    if validate_input():
        values = tree.item(selected, "values")
        record_id = values[0]  
        
        first = fname_entry.get().strip()
        middle = mname_entry.get().strip()
        last = lname_entry.get().strip()
        by = int(birth_entry.get().strip())
        age = 2026 - by
        
        workbook = op.load_workbook(EXCEL_FILE)
        sheet = workbook.active
        
        for row in sheet.iter_rows(min_row=2):
            if str(row[0].value) == str(record_id):
                row[1].value = last
                row[2].value = first
                row[3].value = middle
                row[4].value = by
                row[5].value = age
                break
                
        workbook.save(EXCEL_FILE)
        messagebox.showinfo("Success", "Record updated successfully!!")
        display_excel()
        clear_entries()

def delete_data():
    selected = tree.focus()
    if not selected:
        messagebox.showerror("Error", "Select a record first!")
        return
        
    values = tree.item(selected, "values")
    record_id = values[0]
    
    confirm = messagebox.askyesno("confirm", "do you sure you want to delete this record?")
    if confirm:
        workbook = op.load_workbook(EXCEL_FILE)
        sheet = workbook.active
        
        for idx, row in enumerate(sheet.iter_rows(min_row=2, values_only=True), start=2):
            if str(row[0]) == str(record_id):
                sheet.delete_rows(idx)
                break
                
        workbook.save(EXCEL_FILE)
        messagebox.showinfo("Success", "Record deleted successfully!!")
        display_excel()
        clear_entries()

def clear_entries():
    """Clears all textboxes in the form"""
    lname_entry.delete(0, tk.END)
    fname_entry.delete(0, tk.END)
    mname_entry.delete(0, tk.END)
    birth_entry.delete(0, tk.END)


window = tk.Tk()
window.title("Customer Information System")
window.geometry("820x450")
window.configure(bg="lightgreen")

title_lbl = tk.Label(window, text="Customer Profile Builder", font=("Times New Roman", 14, "bold"), bg="lightgreen")
title_lbl.pack(pady=10)

genframe = tk.Frame(window, bg="lightgreen", bd=2, relief="groove")
genframe.pack(side="left", padx=15, pady=10, fill="y")

tk.Label(genframe, text="First Name:", bg="lightgreen", font=("Poppins", 10, "italic")).grid(row=0, column=0, sticky="w", pady=5, padx=5)
fname_entry = tk.Entry(genframe, font=("Poppins", 12))
fname_entry.grid(row=0, column=1, pady=5, padx=5)

tk.Label(genframe, text="Middle Name:", bg="lightgreen", font=("Poppins", 10, "italic")).grid(row=1, column=0, sticky="w", pady=5, padx=5)
mname_entry = tk.Entry(genframe, font=("Poppins", 12))
mname_entry.grid(row=1, column=1, pady=5, padx=5)

tk.Label(genframe, text="Last Name:", bg="lightgreen", font=("Poppins", 10, "italic")).grid(row=2, column=0, sticky="w", pady=5, padx=5)
lname_entry = tk.Entry(genframe, font=("Poppins", 12))
lname_entry.grid(row=2, column=1, pady=5, padx=5)

tk.Label(genframe, text="Birth Year:", bg="lightgreen", font=("Poppins", 10, "italic")).grid(row=3, column=0, sticky="w", pady=5, padx=5)
birth_entry = tk.Entry(genframe, font=("Poppins", 12))
birth_entry.grid(row=3, column=1, pady=5, padx=5)


button_frame = tk.Frame(genframe, bg="lightgreen")
button_frame.grid(row=4, column=0, columnspan=2, pady=15)

submit_btn = tk.Button(button_frame, text="Submit", font=("Poppins", 10, "bold"), bg="lightpink", command=append_excel, width=9)
submit_btn.grid(row=0, column=0, padx=5, pady=5)

update_btn = tk.Button(button_frame, text="Update", font=("Poppins", 10, "bold"), bg="orange", command=update_data, width=9)
update_btn.grid(row=0, column=1, padx=5, pady=5)

delete_btn = tk.Button(button_frame, text="Delete", font=("Poppins", 10, "bold"), bg="red", fg="white", command=delete_data, width=9)
delete_btn.grid(row=1, column=0, padx=5, pady=5)

clear_btn = tk.Button(button_frame, text="Clear", font=("Poppins", 10, "bold"), bg="lightgray", command=clear_entries, width=9)
clear_btn.grid(row=1, column=1, padx=5, pady=5)


table_frame = tk.Frame(window)
table_frame.pack(side="right", fill="both", expand=True, padx=15, pady=10)

cols = ("ID", "Last", "First", "Middle", "BirthYear", "Age")
tree = ttk.Treeview(table_frame, columns=cols, show="headings")


for col in cols:
    tree.heading(col, text=col)
    tree.column(col, width=80, anchor="center")


scrollbar = ttk.Scrollbar(table_frame, orient="vertical", command=tree.yview)
tree.configure(yscrollcommand=scrollbar.set)
scrollbar.pack(side="right", fill="y")
tree.pack(fill="both", expand=True)


tree.bind("<<TreeviewSelect>>", select_record)

display_excel()

window.mainloop()
